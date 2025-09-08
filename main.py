import os
import re
from collections import Counter
from typing import List, Dict, Tuple, Optional, Union
from pathlib import Path
import logging
from datetime import datetime
from urllib.parse import urlparse

# Importa o gerenciador de stop words
from config import StopWordsManager

# Configuração de logging para debug e monitoramento
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class AnalisadorDocumental:
    """
    Classe para análise de documentos PDF e páginas web com funcionalidades de processamento de texto
    e integração com API da OpenAI.
    
    Atributos:
        fonte (str): Caminho para arquivo PDF ou URL da página web
        tipo_fonte (str): Tipo da fonte ('pdf' ou 'web')
        conteudo (str): Conteúdo extraído da fonte
        palavras_processadas (List[str]): Lista de palavras após limpeza
        api_key (str): Chave da API da OpenAI
        info_fonte (Dict): Informações gerais da fonte
        stop_words_manager (StopWordsManager): Gerenciador de stop words
        idioma_detectado (str): Idioma detectado do conteúdo
    """
    
    def __init__(self, fonte: str, api_key: Optional[str] = None, idioma: Optional[str] = None):
        """
        Inicializa o analisador de documentos.
        
        Args:
            fonte (str): Caminho para arquivo PDF ou URL da página web
            api_key (str, optional): Chave da API da OpenAI para análise com ChatGPT
            idioma (str, optional): Idioma para stop words ('portugues', 'ingles', 'espanhol')
        """
        self.fonte = fonte
        self.tipo_fonte = self._determinar_tipo_fonte(fonte)
        self.conteudo = ""
        self.palavras_processadas = []
        self.api_key = api_key or os.getenv('OPENAI_API_KEY')
        self.info_fonte = {}
        
        # Inicializa gerenciador de stop words
        self.stop_words_manager = StopWordsManager()
        self.idioma_detectado = idioma
        
        # Validação inicial
        self._validar_fonte()
        self._extrair_informacoes_fonte()
        self._extrair_conteudo()
        self._processar_palavras()
    
    def _determinar_tipo_fonte(self, fonte: str) -> str:
        """
        Determina se a fonte é um arquivo PDF ou uma URL web.
        
        Args:
            fonte (str): Caminho do arquivo ou URL
            
        Returns:
            str: 'pdf' ou 'web'
        """
        # Verifica se é uma URL
        try:
            resultado = urlparse(fonte)
            if resultado.scheme and resultado.netloc:
                return 'web'
        except:
            pass
        
        # Verifica se é um arquivo PDF
        if fonte.lower().endswith('.pdf'):
            return 'pdf'
        
        # Verifica se é um caminho de arquivo existente
        if os.path.exists(fonte):
            return 'pdf'
        
        # Se não conseguir determinar, assume que é web
        return 'web'
    
    def _normalizar_caminho(self, caminho: str) -> str:
        """
        Normaliza o caminho do arquivo para evitar problemas de encoding no Windows.
        
        Args:
            caminho (str): Caminho original do arquivo
            
        Returns:
            str: Caminho normalizado
        """
        try:
            # Usa pathlib para normalizar o caminho de forma segura
            path_obj = Path(caminho)
            return str(path_obj.resolve())
        except Exception as e:
            logger.warning(f"Erro ao normalizar caminho: {e}. Usando caminho original.")
            return caminho
    
    def _validar_fonte(self) -> None:
        """
        Valida se a fonte existe e é válida.
        
        Raises:
            FileNotFoundError: Se o arquivo não existir
            ValueError: Se a fonte não for válida
        """
        try:
            if self.tipo_fonte == 'pdf':
                if not os.path.exists(self.fonte):
                    raise FileNotFoundError(f"Arquivo não encontrado: {self.fonte}")
                
                if not self.fonte.lower().endswith('.pdf'):
                    raise ValueError("O arquivo deve ser um PDF")
            
            elif self.tipo_fonte == 'web':
                # Validação básica de URL
                try:
                    resultado = urlparse(self.fonte)
                    if not resultado.scheme or not resultado.netloc:
                        raise ValueError("URL inválida")
                except Exception as e:
                    raise ValueError(f"URL inválida: {e}")
            
            logger.info(f"Fonte validada: {self.fonte} (tipo: {self.tipo_fonte})")
            
        except Exception as e:
            logger.error(f"Erro na validação da fonte: {e}")
            raise
    
    def _extrair_informacoes_fonte(self) -> None:
        """
        Extrai informações gerais da fonte (PDF ou web).
        """
        try:
            if self.tipo_fonte == 'pdf':
                self._extrair_informacoes_pdf()
            elif self.tipo_fonte == 'web':
                self._extrair_informacoes_web()
            
            logger.info(f"Informações da fonte extraídas")
            
        except Exception as e:
            logger.error(f"Erro ao extrair informações da fonte: {e}")
    
    def _extrair_informacoes_pdf(self) -> None:
        """
        Extrai informações de um arquivo PDF.
        """
        # Informações básicas do arquivo (sempre disponíveis)
        self.info_fonte = {
            'tipo': 'PDF',
            'nome': os.path.basename(self.fonte),
            'caminho_completo': self.fonte,
            'tamanho_bytes': os.path.getsize(self.fonte),
            'data_modificacao': datetime.fromtimestamp(
                os.path.getmtime(self.fonte)
            ).strftime('%d/%m/%Y %H:%M:%S'),
            'total_paginas': 'Não disponível (bibliotecas não instaladas)',
            'versao_pdf': 'Não disponível',
            'criptografado': 'Não disponível',
            'metadados': {}
        }
        
        # Tenta extrair informações usando diferentes bibliotecas
        self._tentar_extrair_metadados()
    
    def _extrair_informacoes_web(self) -> None:
        """
        Extrai informações de uma página web.
        """
        try:
            import requests
            from bs4 import BeautifulSoup
            
            # Faz requisição para obter informações básicas
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(self.fonte, headers=headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extrai informações da página
            titulo = soup.find('title')
            titulo_texto = titulo.get_text().strip() if titulo else 'Sem título'
            
            # Extrai meta tags
            meta_tags = {}
            for meta in soup.find_all('meta'):
                name = meta.get('name') or meta.get('property')
                content = meta.get('content')
                if name and content:
                    meta_tags[name] = content
            
            self.info_fonte = {
                'tipo': 'Página Web',
                'url': self.fonte,
                'titulo': titulo_texto,
                'tamanho_bytes': len(response.content),
                'data_acesso': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'status_code': response.status_code,
                'content_type': response.headers.get('content-type', 'Desconhecido'),
                'encoding': response.encoding,
                'metadados': meta_tags
            }
            
        except ImportError:
            logger.warning("requests ou beautifulsoup4 não instalados. Usando informações básicas.")
            self.info_fonte = {
                'tipo': 'Página Web',
                'url': self.fonte,
                'titulo': 'Não disponível (bibliotecas não instaladas)',
                'tamanho_bytes': 0,
                'data_acesso': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'status_code': 'Não disponível',
                'content_type': 'Não disponível',
                'encoding': 'Não disponível',
                'metadados': {}
            }
        except Exception as e:
            logger.error(f"Erro ao extrair informações da web: {e}")
            self.info_fonte = {
                'tipo': 'Página Web',
                'url': self.fonte,
                'titulo': f'Erro: {str(e)}',
                'tamanho_bytes': 0,
                'data_acesso': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'status_code': 'Erro',
                'content_type': 'Erro',
                'encoding': 'Erro',
                'metadados': {}
            }
    
    def _tentar_extrair_metadados(self) -> None:
        """
        Tenta extrair metadados usando diferentes bibliotecas disponíveis.
        """
        # Tenta com PyMuPDF (fitz) primeiro
        if self._extrair_metadados_pymupdf():
            return
        
        # Tenta com pdfplumber
        if self._extrair_metadados_pdfplumber():
            return
        
        logger.info("Nenhuma biblioteca de PDF disponível para extrair metadados")
    
    def _extrair_metadados_pymupdf(self) -> bool:
        """
        Tenta extrair metadados usando PyMuPDF (fitz).
        
        Returns:
            bool: True se conseguiu extrair, False caso contrário
        """
        try:
            import fitz  # PyMuPDF
            
            doc = fitz.open(self.fonte)
            
            # Informações básicas
            self.info_fonte.update({
                'total_paginas': len(doc),
                'versao_pdf': f"{doc.version}",
                'criptografado': doc.needs_pass,
            })
            
            # Metadados
            metadata = doc.metadata
            if metadata:
                self.info_fonte['metadados'] = {
                    'titulo': metadata.get('title', 'Não informado'),
                    'autor': metadata.get('author', 'Não informado'),
                    'assunto': metadata.get('subject', 'Não informado'),
                    'criador': metadata.get('creator', 'Não informado'),
                    'produtor': metadata.get('producer', 'Não informado'),
                    'data_criacao': metadata.get('creationDate', 'Não informado'),
                    'data_modificacao_meta': metadata.get('modDate', 'Não informado')
                }
            
            doc.close()
            logger.info(f"Metadados PyMuPDF extraídos: {self.info_fonte['total_paginas']} páginas")
            return True
            
        except ImportError:
            logger.info("PyMuPDF não está instalado")
            return False
        except Exception as e:
            logger.warning(f"Erro ao extrair metadados com PyMuPDF: {e}")
            return False
    
    def _extrair_metadados_pdfplumber(self) -> bool:
        """
        Tenta extrair metadados usando pdfplumber.
        
        Returns:
            bool: True se conseguiu extrair, False caso contrário
        """
        try:
            import pdfplumber
            
            with pdfplumber.open(self.fonte) as pdf:
                # Informações básicas
                self.info_fonte.update({
                    'total_paginas': len(pdf.pages),
                    'versao_pdf': 'PDF (pdfplumber)',
                    'criptografado': False,  # pdfplumber não detecta criptografia facilmente
                })
                
                # Metadados
                if hasattr(pdf, 'metadata') and pdf.metadata:
                    metadata = pdf.metadata
                    self.info_fonte['metadados'] = {
                        'titulo': metadata.get('Title', 'Não informado'),
                        'autor': metadata.get('Author', 'Não informado'),
                        'assunto': metadata.get('Subject', 'Não informado'),
                        'criador': metadata.get('Creator', 'Não informado'),
                        'produtor': metadata.get('Producer', 'Não informado'),
                        'data_criacao': metadata.get('CreationDate', 'Não informado'),
                        'data_modificacao_meta': metadata.get('ModDate', 'Não informado')
                    }
                
                logger.info(f"Metadados pdfplumber extraídos: {self.info_fonte['total_paginas']} páginas")
                return True
                
        except ImportError:
            logger.info("pdfplumber não está instalado")
            return False
        except Exception as e:
            logger.warning(f"Erro ao extrair metadados com pdfplumber: {e}")
            return False
    
    
    def _extrair_conteudo(self) -> None:
        """
        Extrai o conteúdo da fonte (PDF ou web).
        """
        try:
            if self.tipo_fonte == 'pdf':
                self._extrair_conteudo_pdf()
            elif self.tipo_fonte == 'web':
                self._extrair_conteudo_web()
            
        except Exception as e:
            logger.error(f"Erro ao extrair conteúdo: {e}")
            self._usar_conteudo_simulado()
    
    def _extrair_conteudo_pdf(self) -> None:
        """
        Extrai o conteúdo de um arquivo PDF.
        """
        # Tenta diferentes métodos de extração em ordem de preferência
        conteudo = self._extrair_com_pymupdf()
        if conteudo:
            self.conteudo = conteudo
            return
        
        conteudo = self._extrair_com_pdfplumber()
        if conteudo:
            self.conteudo = conteudo
            return
        
        # Fallback: conteúdo simulado para demonstração
        self._usar_conteudo_simulado()
    
    def _extrair_conteudo_web(self) -> None:
        """
        Extrai o conteúdo de uma página web.
        """
        try:
            import requests
            from bs4 import BeautifulSoup
            
            # Configuração da requisição
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            logger.info(f"Fazendo requisição para: {self.fonte}")
            response = requests.get(self.fonte, headers=headers, timeout=30)
            response.raise_for_status()
            
            # Parse do HTML
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Remove scripts e estilos
            for script in soup(["script", "style", "nav", "footer", "header"]):
                script.decompose()
            
            # Extrai texto
            texto = soup.get_text()
            
            # Limpa o texto
            linhas = (linha.strip() for linha in texto.splitlines())
            chunks = (frase.strip() for linha in linhas for frase in linha.split("  "))
            texto_limpo = ' '.join(chunk for chunk in chunks if chunk)
            
            if texto_limpo:
                self.conteudo = texto_limpo
                logger.info(f"Conteúdo web extraído: {len(self.conteudo)} caracteres")
            else:
                logger.warning("Não foi possível extrair texto da página web")
                self._usar_conteudo_simulado()
                
        except ImportError:
            logger.error("requests ou beautifulsoup4 não instalados. Instale com: pip install requests beautifulsoup4")
            self._usar_conteudo_simulado()
        except Exception as e:
            logger.error(f"Erro ao extrair conteúdo web: {e}")
            self._usar_conteudo_simulado()
    
    def _extrair_com_pymupdf(self) -> Optional[str]:
        """
        Tenta extrair conteúdo usando PyMuPDF (fitz).
        
        Returns:
            Optional[str]: Conteúdo extraído ou None se falhar
        """
        try:
            import fitz  # PyMuPDF
            
            doc = fitz.open(self.fonte)
            
            # Verifica se o PDF está criptografado
            if doc.needs_pass:
                logger.warning("PDF está criptografado. Tentando extrair sem senha...")
                try:
                    doc.authenticate("")  # Tenta sem senha
                except:
                    logger.error("PDF criptografado e não foi possível descriptografar")
                    doc.close()
                    return None
            
            # Extrai texto de todas as páginas
            conteudos_paginas = []
            total_paginas = len(doc)
            
            logger.info(f"Iniciando extração com PyMuPDF: {total_paginas} páginas...")
            
            for i in range(total_paginas):
                try:
                    pagina = doc.load_page(i)
                    texto_pagina = pagina.get_text()
                    if texto_pagina:
                        conteudos_paginas.append(texto_pagina)
                    
                    # Log de progresso a cada 10 páginas
                    if (i + 1) % 10 == 0 or (i + 1) == total_paginas:
                        logger.info(f"Página {i + 1}/{total_paginas} processada")
                        
                except Exception as e:
                    logger.warning(f"Erro ao extrair página {i + 1}: {e}")
                    continue
            
            doc.close()
            
            # Junta todo o conteúdo
            conteudo = " ".join(conteudos_paginas)
            conteudo = re.sub(r'\s+', ' ', conteudo).strip()
            
            if conteudo:
                logger.info(f"Conteúdo extraído com PyMuPDF: {len(conteudo)} caracteres")
                return conteudo
            else:
                logger.warning("PyMuPDF não conseguiu extrair texto do PDF")
                return None
                
        except ImportError:
            logger.info("PyMuPDF não está instalado")
            return None
        except Exception as e:
            logger.error(f"Erro ao extrair com PyMuPDF: {e}")
            return None
    
    def _extrair_com_pdfplumber(self) -> Optional[str]:
        """
        Returns:
            Optional[str]: Conteúdo extraído ou None se falhar
        """
        try:
            import pdfplumber
            
            with pdfplumber.open(self.fonte) as pdf:
                conteudos_paginas = []
                total_paginas = len(pdf.pages)
                
                logger.info(f"Iniciando extração com pdfplumber: {total_paginas} páginas...")
                
                for i, pagina in enumerate(pdf.pages, 1):
                    try:
                        texto_pagina = pagina.extract_text()
                        if texto_pagina:
                            conteudos_paginas.append(texto_pagina)
                        
                        # Log de progresso a cada 10 páginas
                        if i % 10 == 0 or i == total_paginas:
                            logger.info(f"Página {i}/{total_paginas} processada")
                            
                    except Exception as e:
                        logger.warning(f"Erro ao extrair página {i}: {e}")
                        continue
                
                # Junta todo o conteúdo
                conteudo = " ".join(conteudos_paginas)
                conteudo = re.sub(r'\s+', ' ', conteudo).strip()
                
                if conteudo:
                    logger.info(f"Conteúdo extraído com pdfplumber: {len(conteudo)} caracteres")
                    return conteudo
                else:
                    logger.warning("pdfplumber não conseguiu extrair texto do PDF")
                    return None
                    
        except ImportError:
            logger.info("pdfplumber não está instalado")
            return None
        except Exception as e:
            logger.error(f"Erro ao extrair com pdfplumber: {e}")
            return None
    
    def _usar_conteudo_simulado(self) -> None:
        """
        Usa conteúdo simulado quando não é possível extrair da fonte real.
        """
        logger.warning("Usando conteúdo simulado - bibliotecas não disponíveis")
        
        if self.tipo_fonte == 'pdf':
            self.conteudo = """
            Este é um documento PDF de exemplo para análise. O documento contém informações importantes
            sobre análise documental e processamento de texto. A análise de frequência de palavras
            é uma técnica fundamental em processamento de linguagem natural.
            
            Palavras como "análise", "documento", "processamento" aparecem frequentemente neste texto.
            O objetivo é demonstrar como funciona a análise de frequência e ranking de palavras.
            
            Para extrair conteúdo real de PDFs, instale uma das bibliotecas:
            - pip install PyMuPDF (recomendado)
            - pip install pdfplumber
            """
        else:
            self.conteudo = """
            Este é um conteúdo de página web de exemplo para análise. A página contém informações importantes
            sobre análise documental e processamento de texto. A análise de frequência de palavras
            é uma técnica fundamental em processamento de linguagem natural.
            
            Palavras como "análise", "documento", "processamento" aparecem frequentemente neste texto.
            O objetivo é demonstrar como funciona a análise de frequência e ranking de palavras.
            
            Para extrair conteúdo real de páginas web, instale:
            - pip install requests beautifulsoup4
            """
        
        logger.info(f"Conteúdo simulado criado: {len(self.conteudo)} caracteres")
    
    def obter_informacoes_fonte(self) -> Dict[str, any]:
        """
        Retorna informações detalhadas sobre a fonte (PDF ou web).
        
        Returns:
            Dict[str, any]: Dicionário com informações da fonte
        """
        return self.info_fonte
    
    def _processar_palavras(self) -> None:
        """
        Processa e limpa as palavras do conteúdo extraído.
        Remove pontuação, converte para minúsculas e filtra palavras comuns.
        """
        # Detecta idioma se não foi especificado
        if not self.idioma_detectado:
            self.idioma_detectado = self.stop_words_manager.detect_language(self.conteudo)
            logger.info(f"Idioma detectado: {self.idioma_detectado}")
        
        # Obtém stop words para o idioma detectado
        stop_words = self.stop_words_manager.get_stop_words(self.idioma_detectado)
        
        # Limpeza e processamento do texto
        texto_limpo = re.sub(r'[^\w\s]', '', self.conteudo.lower())
        palavras = texto_limpo.split()
        
        # Filtra palavras comuns e palavras muito curtas
        self.palavras_processadas = [
            palavra for palavra in palavras 
            if palavra not in stop_words and len(palavra) > 2
        ]
        
        logger.info(f"Processadas {len(self.palavras_processadas)} palavras úteis (idioma: {self.idioma_detectado})")
    
    def analisar_frequencia_palavras(self, top_n: int = 20) -> List[Tuple[str, int]]:
        """
        Analisa as palavras mais frequentes no documento e retorna um ranking.
        
        Args:
            top_n (int): Número de palavras mais frequentes a retornar
            
        Returns:
            List[Tuple[str, int]]: Lista de tuplas (palavra, frequência) ordenada por frequência
            
        Raises:
            ValueError: Se não houver palavras processadas
        """
        if not self.palavras_processadas:
            raise ValueError("Nenhuma palavra processada disponível para análise")
        
        # Conta a frequência das palavras
        contador = Counter(self.palavras_processadas)
        
        # Retorna as top_n palavras mais frequentes
        ranking = contador.most_common(top_n)
        
        logger.info(f"Análise de frequência concluída. Top {top_n} palavras identificadas")
        return ranking
    
    def obter_estatisticas_gerais(self) -> Dict[str, any]:
        """
        Retorna estatísticas gerais sobre o documento.
        
        Returns:
            Dict[str, any]: Dicionário com estatísticas do documento
        """
        total_palavras = len(self.palavras_processadas)
        palavras_unicas = len(set(self.palavras_processadas))
        
        # Calcula densidade de palavras únicas
        densidade = (palavras_unicas / total_palavras * 100) if total_palavras > 0 else 0
        
        estatisticas = {
            'total_palavras': total_palavras,
            'palavras_unicas': palavras_unicas,
            'densidade_vocabulario': round(densidade, 2),
            'tamanho_conteudo': len(self.conteudo),
            'fonte': self.fonte,
            'tipo_fonte': self.tipo_fonte,
            'idioma_detectado': self.idioma_detectado
        }
        
        logger.info("Estatísticas gerais calculadas")
        return estatisticas
    

    
    def gerar_relatorio_completo(self, top_n: int = 25) -> str:
        """
        Gera um relatório completo com todas as análises disponíveis.
        
        Args:
            top_n (int): Número de palavras mais frequentes para incluir
            
        Returns:
            str: Relatório formatado com todas as análises
        """
        relatorio = []
        relatorio.append("=" * 80)
        relatorio.append("RELATÓRIO DE ANÁLISE DOCUMENTAL")
        relatorio.append("=" * 80)
        
        # Informações da fonte
        info = self.obter_informacoes_fonte()
        relatorio.append("INFORMAÇÕES DA FONTE:")
        
        if self.tipo_fonte == 'pdf':
            relatorio.append(f"- Tipo: {info['tipo']}")
            relatorio.append(f"- Nome: {info['nome']}")
            relatorio.append(f"- Caminho: {info['caminho_completo']}")
            relatorio.append(f"- Tamanho: {self._formatar_tamanho(info['tamanho_bytes'])}")
            relatorio.append(f"- Data de modificação: {info['data_modificacao']}")
            relatorio.append(f"- Total de páginas: {info['total_paginas']}")
            relatorio.append(f"- Versão PDF: {info['versao_pdf']}")
            relatorio.append(f"- Criptografado: {'Sim' if info['criptografado'] else 'Não'}")
        else:
            relatorio.append(f"- Tipo: {info['tipo']}")
            relatorio.append(f"- URL: {info['url']}")
            relatorio.append(f"- Título: {info['titulo']}")
            relatorio.append(f"- Tamanho: {self._formatar_tamanho(info['tamanho_bytes'])}")
            relatorio.append(f"- Data de acesso: {info['data_acesso']}")
            relatorio.append(f"- Status: {info['status_code']}")
            relatorio.append(f"- Content-Type: {info['content_type']}")
            relatorio.append(f"- Encoding: {info['encoding']}")
        
        # Metadados se disponíveis
        if info['metadados']:
            relatorio.append("")
            relatorio.append("METADADOS:")
            for chave, valor in info['metadados'].items():
                if valor and valor != 'Não informado':
                    relatorio.append(f"- {chave.replace('_', ' ').title()}: {valor}")
        
        relatorio.append("")
        
        # Estatísticas gerais
        stats = self.obter_estatisticas_gerais()
        relatorio.append("ESTATÍSTICAS GERAIS:")
        relatorio.append(f"- Total de palavras: {stats['total_palavras']:,}")
        relatorio.append(f"- Palavras únicas: {stats['palavras_unicas']:,}")
        relatorio.append(f"- Densidade de vocabulário: {stats['densidade_vocabulario']}%")
        relatorio.append(f"- Tamanho do conteúdo: {stats['tamanho_conteudo']:,} caracteres")
        relatorio.append(f"- Idioma detectado: {stats['idioma_detectado']}")
        relatorio.append("")
        
        # Ranking de palavras mais frequentes
        relatorio.append(f"TOP {top_n} PALAVRAS MAIS FREQUENTES:")
        ranking = self.analisar_frequencia_palavras(top_n)
        for i, (palavra, frequencia) in enumerate(ranking, 1):
            relatorio.append(f"{i:2d}. {palavra:<25} - {frequencia:5d} ocorrências")
        
        relatorio.append("")
        relatorio.append("=" * 80)
        
        return "\n".join(relatorio)
    
    def _formatar_tamanho(self, bytes_size: int) -> str:
        """
        Formata o tamanho em bytes para uma representação legível.
        
        Args:
            bytes_size (int): Tamanho em bytes
            
        Returns:
            str: Tamanho formatado (ex: "1.5 MB")
        """
        for unidade in ['B', 'KB', 'MB', 'GB']:
            if bytes_size < 1024.0:
                return f"{bytes_size:.1f} {unidade}"
            bytes_size /= 1024.0
        return f"{bytes_size:.1f} TB"


class AnalisadorMultiplosDocumentos:
    """
    Classe para análise de múltiplos documentos PDF em uma pasta.
    Realiza análise individual de cada arquivo e gera um ranking geral consolidado.
    
    Atributos:
        pasta (str): Caminho para a pasta contendo os PDFs
        analisadores (List[AnalisadorDocumental]): Lista de analisadores para cada documento
        resultados_gerais (Dict): Resultados consolidados de todos os documentos
    """
    
    def __init__(self, pasta: str, api_key: Optional[str] = None, idioma: Optional[str] = None):
        """
        Inicializa o analisador de múltiplos documentos.
        
        Args:
            pasta (str): Caminho para a pasta contendo os PDFs
            api_key (str, optional): Chave da API da OpenAI para análise com ChatGPT
            idioma (str, optional): Idioma para stop words ('portugues', 'ingles', 'espanhol')
        """
        self.pasta = pasta
        self.api_key = api_key
        self.idioma = idioma
        self.analisadores = []
        self.resultados_gerais = {}
        
        # Valida a pasta
        self._validar_pasta()
        
        # Encontra e analisa todos os PDFs
        self._encontrar_e_analisar_pdfs()
        
        # Gera resultados consolidados
        self._gerar_resultados_consolidados()
    
    def _validar_pasta(self) -> None:
        """
        Valida se a pasta existe e é acessível.
        
        Raises:
            FileNotFoundError: Se a pasta não existir
            ValueError: Se o caminho não for uma pasta
        """
        if not os.path.exists(self.pasta):
            raise FileNotFoundError(f"Pasta não encontrada: {self.pasta}")
        
        if not os.path.isdir(self.pasta):
            raise ValueError(f"O caminho não é uma pasta: {self.pasta}")
        
        logger.info(f"Pasta validada: {self.pasta}")
    
    def _encontrar_pdfs(self) -> List[str]:
        """
        Encontra todos os arquivos PDF na pasta especificada.
        
        Returns:
            List[str]: Lista de caminhos completos para os PDFs encontrados
        """
        pdfs = []
        
        try:
            for arquivo in os.listdir(self.pasta):
                if arquivo.lower().endswith('.pdf'):
                    caminho_completo = os.path.join(self.pasta, arquivo)
                    pdfs.append(caminho_completo)
            
            logger.info(f"Encontrados {len(pdfs)} arquivos PDF na pasta")
            return pdfs
            
        except Exception as e:
            logger.error(f"Erro ao buscar PDFs na pasta: {e}")
            return []
    
    def _encontrar_e_analisar_pdfs(self) -> None:
        """
        Encontra todos os PDFs na pasta e cria analisadores para cada um.
        """
        pdfs = self._encontrar_pdfs()
        
        if not pdfs:
            logger.warning("Nenhum arquivo PDF encontrado na pasta")
            return
        
        logger.info(f"Iniciando análise de {len(pdfs)} documentos...")
        
        for i, pdf in enumerate(pdfs, 1):
            try:
                logger.info(f"Analisando documento {i}/{len(pdfs)}: {os.path.basename(pdf)}")
                analisador = AnalisadorDocumental(pdf, self.api_key, self.idioma)
                self.analisadores.append(analisador)
                
            except Exception as e:
                logger.error(f"Erro ao analisar {pdf}: {e}")
                continue
        
        logger.info(f"Análise concluída: {len(self.analisadores)} documentos processados com sucesso")
    
    def _gerar_resultados_consolidados(self) -> None:
        """
        Gera resultados consolidados de todos os documentos analisados.
        """
        if not self.analisadores:
            logger.warning("Nenhum analisador disponível para consolidar resultados")
            return
        
        # Estatísticas gerais
        total_documentos = len(self.analisadores)
        total_palavras_geral = 0
        total_palavras_unicas_geral = 0
        total_tamanho_conteudo = 0
        
        # Ranking consolidado de palavras
        ranking_consolidado = Counter()
        
        # Informações individuais de cada documento
        documentos_info = []
        
        for analisador in self.analisadores:
            try:
                # Estatísticas do documento
                stats = analisador.obter_estatisticas_gerais()
                info_fonte = analisador.obter_informacoes_fonte()
                
                # Acumula estatísticas gerais
                total_palavras_geral += stats['total_palavras']
                total_palavras_unicas_geral += stats['palavras_unicas']
                total_tamanho_conteudo += stats['tamanho_conteudo']
                
                # Acumula ranking de palavras
                ranking_doc = analisador.analisar_frequencia_palavras(top_n=100)
                for palavra, frequencia in ranking_doc:
                    ranking_consolidado[palavra] += frequencia
                
                # Informações do documento
                doc_info = {
                    'nome': info_fonte.get('nome', 'Desconhecido'),
                    'caminho': info_fonte.get('caminho_completo', 'Desconhecido'),
                    'total_palavras': stats['total_palavras'],
                    'palavras_unicas': stats['palavras_unicas'],
                    'densidade_vocabulario': stats['densidade_vocabulario'],
                    'tamanho_conteudo': stats['tamanho_conteudo'],
                    'idioma': stats['idioma_detectado'],
                    'total_paginas': info_fonte.get('total_paginas', 'Desconhecido'),
                    'tamanho_arquivo': info_fonte.get('tamanho_bytes', 0)
                }
                documentos_info.append(doc_info)
                
            except Exception as e:
                logger.error(f"Erro ao processar estatísticas do documento: {e}")
                continue
        
        # Calcula estatísticas consolidadas
        densidade_geral = (total_palavras_unicas_geral / total_palavras_geral * 100) if total_palavras_geral > 0 else 0
        
        self.resultados_gerais = {
            'total_documentos': total_documentos,
            'total_palavras': total_palavras_geral,
            'total_palavras_unicas': total_palavras_unicas_geral,
            'densidade_vocabulario_geral': round(densidade_geral, 2),
            'tamanho_conteudo_total': total_tamanho_conteudo,
            'ranking_consolidado': ranking_consolidado.most_common(50),
            'documentos': documentos_info,
            'pasta_analisada': self.pasta,
            'data_analise': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        
        logger.info("Resultados consolidados gerados com sucesso")
    
    def obter_ranking_geral(self, top_n: int = 20) -> List[Tuple[str, int]]:
        """
        Retorna o ranking geral das palavras mais frequentes em todos os documentos.
        
        Args:
            top_n (int): Número de palavras mais frequentes a retornar
            
        Returns:
            List[Tuple[str, int]]: Lista de tuplas (palavra, frequência) ordenada por frequência
        """
        if not self.resultados_gerais:
            return []
        
        return self.resultados_gerais['ranking_consolidado'][:top_n]
    
    def obter_estatisticas_gerais(self) -> Dict[str, any]:
        """
        Retorna estatísticas gerais consolidadas de todos os documentos.
        
        Returns:
            Dict[str, any]: Dicionário com estatísticas consolidadas
        """
        if not self.resultados_gerais:
            return {}
        
        return {
            'total_documentos': self.resultados_gerais['total_documentos'],
            'total_palavras': self.resultados_gerais['total_palavras'],
            'total_palavras_unicas': self.resultados_gerais['total_palavras_unicas'],
            'densidade_vocabulario_geral': self.resultados_gerais['densidade_vocabulario_geral'],
            'tamanho_conteudo_total': self.resultados_gerais['tamanho_conteudo_total'],
            'pasta_analisada': self.resultados_gerais['pasta_analisada'],
            'data_analise': self.resultados_gerais['data_analise']
        }
    
    def obter_documentos_analisados(self) -> List[Dict[str, any]]:
        """
        Retorna informações detalhadas de cada documento analisado.
        
        Returns:
            List[Dict[str, any]]: Lista com informações de cada documento
        """
        if not self.resultados_gerais:
            return []
        
        return self.resultados_gerais['documentos']
    
    def obter_ranking_individual(self, indice_documento: int, top_n: int = 20) -> List[Tuple[str, int]]:
        """
        Retorna o ranking de palavras de um documento específico.
        
        Args:
            indice_documento (int): Índice do documento na lista de analisadores (baseado em 0)
            top_n (int): Número de palavras mais frequentes a retornar
            
        Returns:
            List[Tuple[str, int]]: Lista de tuplas (palavra, frequência) ordenada por frequência
        """
        if not self.analisadores or indice_documento >= len(self.analisadores):
            return []
        
        try:
            analisador = self.analisadores[indice_documento]
            ranking = analisador.analisar_frequencia_palavras(top_n)
            return ranking
        except Exception as e:
            logger.error(f"Erro ao obter ranking individual do documento {indice_documento}: {e}")
            return []
    
    def _obter_ranking_individual(self, indice_documento: int, top_n: int = 20) -> List[Tuple[str, int]]:
        """
        Método interno para obter ranking individual (chama o método público).
        
        Args:
            indice_documento (int): Índice do documento na lista de analisadores
            top_n (int): Número de palavras mais frequentes a retornar
            
        Returns:
            List[Tuple[str, int]]: Lista de tuplas (palavra, frequência) ordenada por frequência
        """
        return self.obter_ranking_individual(indice_documento, top_n)
    
    def gerar_relatorio_completo(self, top_n: int = 25) -> str:
        """
        Gera um relatório completo com análise de todos os documentos.
        
        Args:
            top_n (int): Número de palavras mais frequentes para incluir no ranking geral
            
        Returns:
            str: Relatório formatado com todas as análises
        """
        if not self.resultados_gerais:
            return "Nenhum documento foi analisado com sucesso."
        
        relatorio = []
        relatorio.append("=" * 100)
        relatorio.append("RELATÓRIO DE ANÁLISE DOCUMENTAL - MÚLTIPLOS DOCUMENTOS")
        relatorio.append("=" * 100)
        
        # Informações gerais
        stats = self.obter_estatisticas_gerais()
        relatorio.append("INFORMAÇÕES GERAIS:")
        relatorio.append(f"- Pasta analisada: {stats['pasta_analisada']}")
        relatorio.append(f"- Data da análise: {stats['data_analise']}")
        relatorio.append(f"- Total de documentos: {stats['total_documentos']}")
        relatorio.append(f"- Total de palavras: {stats['total_palavras']:,}")
        relatorio.append(f"- Total de palavras únicas: {stats['total_palavras_unicas']:,}")
        relatorio.append(f"- Densidade de vocabulário geral: {stats['densidade_vocabulario_geral']}%")
        relatorio.append(f"- Tamanho total do conteúdo: {stats['tamanho_conteudo_total']:,} caracteres")
        relatorio.append("")
        
        # Ranking geral consolidado
        relatorio.append(f"RANKING GERAL - TOP {top_n} PALAVRAS MAIS FREQUENTES:")
        ranking_geral = self.obter_ranking_geral(top_n)
        for i, (palavra, frequencia) in enumerate(ranking_geral, 1):
            relatorio.append(f"{i:2d}. {palavra:<25} - {frequencia:5d} ocorrências")
        
        relatorio.append("")
        
        # Detalhes de cada documento
        relatorio.append("DETALHES DOS DOCUMENTOS ANALISADOS:")
        relatorio.append("-" * 100)
        
        documentos = self.obter_documentos_analisados()
        for i, doc in enumerate(documentos, 1):
            relatorio.append(f"")
            relatorio.append(f"DOCUMENTO {i}: {doc['nome']}")
            relatorio.append(f"  - Caminho: {doc['caminho']}")
            relatorio.append(f"  - Total de palavras: {doc['total_palavras']:,}")
            relatorio.append(f"  - Palavras únicas: {doc['palavras_unicas']:,}")
            relatorio.append(f"  - Densidade de vocabulário: {doc['densidade_vocabulario']}%")
            relatorio.append(f"  - Tamanho do conteúdo: {doc['tamanho_conteudo']:,} caracteres")
            relatorio.append(f"  - Idioma: {doc['idioma']}")
            relatorio.append(f"  - Total de páginas: {doc['total_paginas']}")
            relatorio.append(f"  - Tamanho do arquivo: {self._formatar_tamanho(doc['tamanho_arquivo'])}")
            
            # Ranking individual do documento
            relatorio.append(f"")
            relatorio.append(f"  TOP 20 PALAVRAS MAIS FREQUENTES (DOCUMENTO {i}):")
            ranking_individual = self._obter_ranking_individual(i-1, 20)  # i-1 porque é índice baseado em 0
            if ranking_individual:
                for j, (palavra, frequencia) in enumerate(ranking_individual, 1):
                    relatorio.append(f"    {j:2d}. {palavra:<20} - {frequencia:4d} ocorrências")
            else:
                relatorio.append(f"    Nenhuma palavra encontrada para este documento.")
        
        relatorio.append("")
        relatorio.append("=" * 100)
        
        return "\n".join(relatorio)
    
    def _formatar_tamanho(self, bytes_size: int) -> str:
        """
        Formata o tamanho em bytes para uma representação legível.
        
        Args:
            bytes_size (int): Tamanho em bytes
            
        Returns:
            str: Tamanho formatado (ex: "1.5 MB")
        """
        for unidade in ['B', 'KB', 'MB', 'GB']:
            if bytes_size < 1024.0:
                return f"{bytes_size:.1f} {unidade}"
            bytes_size /= 1024.0
        return f"{bytes_size:.1f} TB"
    
    def verificar_openpyxl_disponivel(self) -> bool:
        """
        Verifica se o openpyxl está disponível e funcionando.
        
        Returns:
            bool: True se openpyxl está disponível, False caso contrário
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            return True
        except ImportError:
            return False
        except Exception:
            return False
    
    def exportar_relatorio_excel(self, caminho_arquivo: str, top_n_palavras: int = 50) -> bool:
        """
        Exporta um relatório completo para Excel com análise de múltiplos documentos.
        Cada linha representa um documento e as colunas contêm as palavras mais frequentes.
        
        Args:
            caminho_arquivo (str): Caminho completo onde salvar o arquivo Excel
            top_n_palavras (int): Número de palavras mais frequentes para incluir como colunas
            
        Returns:
            bool: True se a exportação foi bem-sucedida, False caso contrário
            
        Raises:
            ImportError: Se openpyxl não estiver instalado
            ValueError: Se não houver documentos analisados
        """
        # Verifica se openpyxl está disponível
        if not self.verificar_openpyxl_disponivel():
            raise ImportError(
                "openpyxl não está instalado ou há problema com a instalação. "
                "Execute o script de teste: python test_openpyxl.py "
                "Ou tente reinstalar: pip uninstall openpyxl && pip install openpyxl"
            )
        
        # Importa openpyxl (já verificado que está disponível)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        logger.info("openpyxl importado com sucesso")
        
        if not self.resultados_gerais or not self.analisadores:
            raise ValueError("Nenhum documento foi analisado. Execute a análise primeiro.")
        
        logger.info(f"Iniciando exportação para Excel: {caminho_arquivo}")
        
        # Cria um novo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Análise Documental"
        
        # Estilos para formatação
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Obtém todas as palavras únicas mais frequentes de todos os documentos
        todas_palavras = set()
        for analisador in self.analisadores:
            ranking = analisador.analisar_frequencia_palavras(top_n_palavras)
            for palavra, _ in ranking:
                todas_palavras.add(palavra)
        
        # Ordena as palavras por frequência geral
        palavras_ordenadas = []
        ranking_geral = self.obter_ranking_geral(len(todas_palavras))
        palavras_gerais = {palavra: freq for palavra, freq in ranking_geral}
        
        for palavra in todas_palavras:
            frequencia = palavras_gerais.get(palavra, 0)
            palavras_ordenadas.append((palavra, frequencia))
        
        palavras_ordenadas.sort(key=lambda x: x[1], reverse=True)
        colunas_palavras = [palavra for palavra, _ in palavras_ordenadas[:top_n_palavras]]
        
        # Cabeçalhos das colunas
        colunas = [
            "Documento", "Total Palavras", "Palavras Únicas", "Densidade Vocab.", 
            "Tamanho Conteúdo", "Idioma", "Total Páginas", "Tamanho Arquivo"
        ] + colunas_palavras
        
        # Aplica cabeçalhos
        for col, header in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Preenche os dados de cada documento
        documentos = self.obter_documentos_analisados()
        
        for row, doc in enumerate(documentos, 2):
            # Informações básicas do documento
            ws.cell(row=row, column=1, value=doc['nome']).border = border
            ws.cell(row=row, column=2, value=doc['total_palavras']).border = border
            ws.cell(row=row, column=3, value=doc['palavras_unicas']).border = border
            ws.cell(row=row, column=4, value=f"{doc['densidade_vocabulario']}%").border = border
            ws.cell(row=row, column=5, value=doc['tamanho_conteudo']).border = border
            ws.cell(row=row, column=6, value=doc['idioma']).border = border
            ws.cell(row=row, column=7, value=str(doc['total_paginas'])).border = border
            ws.cell(row=row, column=8, value=self._formatar_tamanho(doc['tamanho_arquivo'])).border = border
            
            # Obtém ranking individual do documento
            indice_doc = row - 2  # Índice baseado em 0
            ranking_individual = self.obter_ranking_individual(indice_doc, top_n_palavras)
            palavras_doc = {palavra: freq for palavra, freq in ranking_individual}
            
            # Preenche frequências das palavras
            for col, palavra in enumerate(colunas_palavras, 9):
                frequencia = palavras_doc.get(palavra, 0)
                cell = ws.cell(row=row, column=col, value=frequencia)
                cell.border = border
                cell.alignment = center_alignment
        
        # Ajusta largura das colunas
        for col in range(1, len(colunas) + 1):
            column_letter = get_column_letter(col)
            if col == 1:  # Coluna do nome do documento
                ws.column_dimensions[column_letter].width = 30
            elif col <= 8:  # Colunas de informações básicas
                ws.column_dimensions[column_letter].width = 15
            else:  # Colunas de palavras
                ws.column_dimensions[column_letter].width = 12
        
        # Adiciona uma segunda aba com resumo geral
        ws_resumo = wb.create_sheet("Resumo Geral")
        
        # Cabeçalho do resumo
        ws_resumo.cell(row=1, column=1, value="Métrica").font = header_font
        ws_resumo.cell(row=1, column=2, value="Valor").font = header_font
        ws_resumo.cell(row=1, column=1).fill = header_fill
        ws_resumo.cell(row=1, column=2).fill = header_fill
        ws_resumo.cell(row=1, column=1).border = border
        ws_resumo.cell(row=1, column=2).border = border
        
        # Dados do resumo
        stats = self.obter_estatisticas_gerais()
        resumo_dados = [
            ("Total de Documentos", stats['total_documentos']),
            ("Total de Palavras", f"{stats['total_palavras']:,}"),
            ("Total de Palavras Únicas", f"{stats['total_palavras_unicas']:,}"),
            ("Densidade de Vocabulário Geral", f"{stats['densidade_vocabulario_geral']}%"),
            ("Tamanho Total do Conteúdo", f"{stats['tamanho_conteudo_total']:,} caracteres"),
            ("Pasta Analisada", stats['pasta_analisada']),
            ("Data da Análise", stats['data_analise'])
        ]
        
        for row, (metrica, valor) in enumerate(resumo_dados, 2):
            ws_resumo.cell(row=row, column=1, value=metrica).border = border
            ws_resumo.cell(row=row, column=2, value=valor).border = border
        
        # Ajusta largura das colunas do resumo
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 50
        
        # Salva o arquivo
        wb.save(caminho_arquivo)
        
        logger.info(f"Relatório Excel exportado com sucesso: {caminho_arquivo}")
        logger.info(f"Arquivo contém {len(documentos)} documentos e {len(colunas_palavras)} palavras analisadas")
        
        return True


def main():
    """
    Função principal para demonstração do uso das classes AnalisadorDocumental e AnalisadorMultiplosDocumentos.
    """
    try:
        print("=== DEMONSTRAÇÃO DO ANALISADOR DOCUMENTAL ===\n")
        
        # Exemplo 1: Análise de PDF individual
        #print("--- ANÁLISE DE PDF INDIVIDUAL ---")
        #analisador_pdf = AnalisadorDocumental("C:/Users/matol/Downloads/PBIA - IA_para_o_Bem_de_Todos.pdf")
        #print(analisador_pdf.gerar_relatorio_completo())
        
        #print("\n" + "="*80 + "\n")
        
        # Exemplo 2: Análise de múltiplos PDFs em uma pasta
        print("--- ANÁLISE DE MÚLTIPLOS PDFs ---")
        pasta_pdfs = "C:/Users/matol/Documents/anpocs2025/corpusdocumental"  # Substitua pelo caminho da sua pasta
        analisador_multiplos = AnalisadorMultiplosDocumentos(pasta_pdfs)
        print(analisador_multiplos.gerar_relatorio_completo())
        
        # Exemplo 3: Exportação para Excel
        print("\n--- EXPORTAÇÃO PARA EXCEL ---")
        try:
            caminho_excel = "relatorio_analise_documental.xlsx"
            sucesso = analisador_multiplos.exportar_relatorio_excel(caminho_excel, top_n_palavras=30)
            if sucesso:
                print(f"✅ Relatório Excel exportado com sucesso: {caminho_excel}")
                print("📊 O arquivo contém:")
                print("   - Aba 'Análise Documental': Cada linha = 1 documento, colunas = palavras mais frequentes")
                print("   - Aba 'Resumo Geral': Estatísticas consolidadas de todos os documentos")
            else:
                print("❌ Erro ao exportar relatório Excel")
        except ImportError as e:
            print(f"⚠️  ERRO DE IMPORTAÇÃO: {e}")
            print("💡 Soluções sugeridas:")
            print("   1. pip uninstall openpyxl && pip install openpyxl")
            print("   2. Execute o script de teste: python test_openpyxl.py")
            print("   3. Verifique se está no ambiente virtual correto")
        except Exception as e:
            print(f"❌ Erro na exportação Excel: {e}")
            print("💡 Execute o script de teste: python test_openpyxl.py")
        
        print("\n" + "="*80 + "\n")
        
        # Exemplo 3: Análise de página web
        #print("--- ANÁLISE DE PÁGINA WEB ---")
        #analisador_web = AnalisadorDocumental("https://www.softwareimprovementgroup.com/us-ai-legislation-overview/")
        #print(analisador_web.gerar_relatorio_completo())
        
        # Exemplo de análise consolidada com ChatGPT (requer API key configurada)
        # api_key = "sua_api_key_aqui"  # Configure sua API key
        # analisador_chatgpt = AnalisadorMultiplosDocumentos(pasta_pdfs, api_key)
        # 
        # prompt = "Analise este conjunto de documentos e identifique os principais temas e tendências."
        # resposta = analisador_chatgpt.analisar_com_chatgpt_geral(prompt)
        # if resposta:
        #     print("\n=== ANÁLISE CONSOLIDADA COM CHATGPT ===")
        #     print(resposta)
        
    except Exception as e:
        print(f"Erro na execução: {e}")


if __name__ == "__main__":
    main()